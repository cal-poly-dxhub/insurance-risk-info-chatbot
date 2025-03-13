import os
from PIL import Image
import pandas as pd
import re
import json
import uuid
from textractor import Textractor
from textractor.visualizers.entitylist import EntityList
from textractor.data.constants import TextractFeatures
import io
import inflect
from collections import OrderedDict
import boto3
import time
import openpyxl
from openpyxl.cell import Cell
from openpyxl.worksheet.cell_range import CellRange
from citation_tools import get_page_num, get_pdf_stream

s3 = boto3.client("s3")
from botocore.config import Config
from langchain_aws import BedrockEmbeddings

config = Config(
    read_timeout=600,
    retries=dict(
        max_attempts=5
    )
)
import re
import pandas as pd
from io import StringIO

bedrock_runtime = boto3.client(service_name='bedrock-runtime', config=config)

INDEX_NAME = "prism-index"


"""
This dictionary `model_dimension_mapping` maps different model names to their respective embedding dimensions.
"""
model_dimension_mapping = {"titanv2": 1024, "titanv1": 1536, "bge": 1024, "all-mini-lm": 384, "e5": 1024}


def _get_emb_(passage, model):
    """
    This function takes a passage of text and a model name as input, and returns the corresponding text embedding.
    The function first checks the provided model name and then invokes the appropriate model or API to generate the text embedding.
    After invoking the appropriate model or API, the function extracts the text embedding from the response and returns it.
    """

    # create an Amazon Titan Text Embeddings client
    embeddings_client = BedrockEmbeddings(model_id="amazon.titan-embed-text-v2:0")

    # Invoke the model
    embedding = embeddings_client.embed_query(passage)
    return (embedding)


def bedrock_streemer(response):
    stream = response.get('body')
    answer = ""
    i = 1
    if stream:
        for event in stream:
            chunk = event.get('chunk')
            if chunk:
                chunk_obj = json.loads(chunk.get('bytes').decode())
                if "delta" in chunk_obj:
                    delta = chunk_obj['delta']
                    if "text" in delta:
                        text = delta['text']
                        #print(text, end="")
                        answer += str(text)
                        i += 1
                if "amazon-bedrock-invocationMetrics" in chunk_obj:
                    input_tokens = chunk_obj['amazon-bedrock-invocationMetrics']['inputTokenCount']
                    output_tokens = chunk_obj['amazon-bedrock-invocationMetrics']['outputTokenCount']
                    print(f"\nInput Tokens: {input_tokens}\nOutput Tokens: {output_tokens}")
    return answer, input_tokens, output_tokens


def bedrock_claude_(chat_history, system_message, prompt, model_id, image_path=None):
    content = []
    if image_path:
        if not isinstance(image_path, list):
            image_path = [image_path]
        for img in image_path:
            s3 = boto3.client('s3')
            match = re.match("s3://(.+?)/(.+)", img)
            image_name = os.path.basename(img)
            _, ext = os.path.splitext(image_name)
            if "jpg" in ext: ext = ".jpeg"
            if match:
                bucket_name = match.group(1)
                key = match.group(2)
                obj = s3.get_object(Bucket=bucket_name, Key=key)
                base_64_encoded_data = base64.b64encode(obj['Body'].read())
                base64_string = base_64_encoded_data.decode('utf-8')
            content.extend([{"type": "text", "text": image_name}, {
                "type": "image",
                "source": {
                    "type": "base64",
                    "media_type": f"image/{ext.lower().replace('.', '')}",
                    "data": base64_string
                }
            }])

    content.append({
        "type": "text",
        "text": prompt
    })
    chat_history.append({"role": "user",
                         "content": content})
    prompt = {
        "anthropic_version": "bedrock-2023-05-31",
        "max_tokens": 1500,
        "temperature": 0.1,
        "system": system_message,
        "messages": chat_history
    }
    answer = ""
    prompt = json.dumps(prompt)
    response = bedrock_runtime.invoke_model_with_response_stream(body=prompt, modelId=model_id,
                                                                 accept="application/json",
                                                                 contentType="application/json")
    answer, input_tokens, output_tokens = bedrock_streemer(response)
    return answer, input_tokens, output_tokens


def _invoke_bedrock_with_retries(current_chat, chat_template, question, model_id, image_path):
    print("Retrying bedrock")
    max_retries = 5
    backoff_base = 2
    max_backoff = 3  # Maximum backoff time in seconds
    retries = 0

    while True:
        try:
            response, input_tokens, output_tokens = bedrock_claude_(current_chat, chat_template, question, model_id,
                                                                    image_path)
            return response, input_tokens, output_tokens
        except ClientError as e:
            if e.response['Error']['Code'] == 'ThrottlingException':
                if retries < max_retries:
                    # Throttling, exponential backoff
                    sleep_time = min(max_backoff, backoff_base ** retries + random.uniform(0, 1))
                    time.sleep(sleep_time)
                    retries += 1
                else:
                    raise e
            elif e.response['Error']['Code'] == 'ModelStreamErrorException':
                if retries < max_retries:
                    # Throttling, exponential backoff
                    sleep_time = min(max_backoff, backoff_base ** retries + random.uniform(0, 1))
                    time.sleep(sleep_time)
                    retries += 1
                else:
                    raise e
            else:
                # Some other API error, rethrow
                raise

def process_document(bucket_name, s3_file, document_url):
    # Textractor
    # BUCKET="YOUR_BUCKET_NAME"
    # file="IRIC-manual-20241-final-draft.pdf" #Change to file path either in S3 or Local

    BUCKET = bucket_name
    file = s3_file

    extractor = Textractor(region_name="us-west-2")

    doc_id = os.path.basename(file)
    file_name, ext = os.path.splitext(file)
    if file.startswith("s3://"):
        document = extractor.start_document_analysis(
            file_source=file,
            features=[TextractFeatures.LAYOUT, TextractFeatures.TABLES],
            # client_request_token=doc_id,
            save_image=False,
            s3_output_path=f"s3://{BUCKET}/textract-output/{file_name}/"

        )
    else:
        document = extractor.start_document_analysis(
            file_source=file,
            features=[TextractFeatures.LAYOUT, TextractFeatures.TABLES],
            # client_request_token=doc_id,
            save_image=False,
            s3_upload_path=f"s3://{BUCKET}",
            s3_output_path=f"s3://{BUCKET}/textract-output/{file_name}/"
        )

    print("Document analysis started... ", end="")

    from textractor.data.text_linearization_config import TextLinearizationConfig

    config = TextLinearizationConfig(
        hide_figure_layout=False,
        title_prefix="<titles><<title>><title>",
        title_suffix="</title><</title>>",
        hide_header_layout=True,
        section_header_prefix="<headers><<header>><header>",
        section_header_suffix="</header><</header>>",
        table_prefix="<tables><table>",
        table_suffix="</table>",
        list_layout_prefix="<<list>><list>",
        list_layout_suffix="</list><</list>>",
        hide_footer_layout=True,
        hide_page_num_layout=True,
    )

    import numpy as np
    def strip_newline(cell):
        """
        A utility function to strip newline characters from a cell.
        Parameters:
        cell (str): The cell value.
        Returns:
        str: The cell value with newline characters removed.
        """
        return str(cell).strip()

    def layout_table_to_excel(document, ids, csv_seperator):
        """
        Converts an Excel table from a document to a Pandas DataFrame,
        handling duplicated values across merged cells.

        Args:
            document: Document containing Excel table
            ids: ID of the Excel table in the document
            csv_seperator: Separator for CSV string conversion

        Returns:
            Pandas DataFrame representation of the Excel table
        """
        # save the table in excel format to preserve the structure of any merged cells
        buffer = io.BytesIO()
        document.tables[ids].to_excel(buffer)
        buffer.seek(0)
        # Load workbook, get active worksheet
        wb = openpyxl.load_workbook(buffer)
        worksheet = wb.active
        # Unmerge cells, duplicate merged values to individual cells
        all_merged_cell_ranges: list[CellRange] = list(
            worksheet.merged_cells.ranges
        )
        for merged_cell_range in all_merged_cell_ranges:
            merged_cell: Cell = merged_cell_range.start_cell
            worksheet.unmerge_cells(range_string=merged_cell_range.coord)
            for row_index, col_index in merged_cell_range.cells:
                cell: Cell = worksheet.cell(row=row_index, column=col_index)
                cell.value = merged_cell.value
        # determine table header index
        df = pd.DataFrame(worksheet.values)
        df = df.map(strip_newline)
        df0 = df.to_csv(sep=csv_seperator, index=False, header=None)
        row_count = len([x for x in df0.split("\n") if x])
        if row_count > 1:
            if not all(value.strip() == '' for value in df0.split("\n")[0].split(csv_seperator)):
                row_count = 1
        # attach table column names
        column_row = 0 if row_count == 1 else 1
        df.columns = df.iloc[column_row]
        df = df[column_row + 1:]
        return df

    def split_list_items_(items):
        """
        Splits the given string into a list of items, handling nested lists.

        Parameters:
        items (str): The input string containing items and possibly nested lists.

        Returns:
        list: A list containing the items extracted from the input string.
        """
        parts = re.split("(<<list>><list>|</list><</list>>)", items)
        output = []

        inside_list = False
        list_item = ""

        for p in parts:
            if p == "<<list>><list>":
                inside_list = True
                list_item = p
            elif p == "</list><</list>>":
                inside_list = False
                list_item += p
                output.append(list_item)
                list_item = ""
            elif inside_list:
                list_item += p.strip()
            else:
                output.extend(p.split('\n'))
        return output

    import io
    """
    This script processes a document containing tables and text. It converts the tables into CSV format 
    and wraps them with XML tags for easy identification. The document structure with text and tables is maintained.
    """
    csv_seperator = "|"  # "\t"
    document_holder = {}
    table_page = {}
    count = 0
    # Whether to handle merged cells by duplicating merged value across corresponding individual cells
    unmerge_span_cells = True
    # Loop through each page in the document
    for ids, page in enumerate(document.pages):
        table_count = len([word for word in page.get_text(config=config).split() if
                           "<tables><table>" in word])  # get the number of table in the extracted document page by header we set earlier
        assert table_count == len(
            page.tables)  # check that number of tables per page is same as *tables extracted by textract TABLE feature
        content = page.get_text(config=config).split("<tables>")
        document_holder[ids] = []
        for idx, item in enumerate(content):
            if "<table>" in item:
                if unmerge_span_cells:
                    df = layout_table_to_excel(document, count, csv_seperator)
                else:
                    df0 = document.tables[count].to_pandas(use_columns=False).to_csv(header=False, index=None,
                                                                                     sep=csv_seperator)
                    row_count = len([x for x in df0.split("\n") if
                                     x])  # Check the number of rows in the parsed table to determine how to read the table headers. if table row count is 1 then headers is obviously at 0 else headers may or may not be at 0
                    # Check if the first row in the csv is empty headers
                    if row_count > 1:
                        if not all(value.strip() == '' for value in df0.split("\n")[0].split(csv_seperator)):
                            row_count = 1
                    df = pd.read_csv(io.StringIO(df0), sep=csv_seperator,
                                     header=0 if row_count == 1 else 1,
                                     keep_default_na=False)  # read table with appropiate column headers
                    df.rename(columns=lambda x: '' if str(x).startswith('Unnamed:') else x, inplace=True)
                table = df.to_csv(index=None, sep=csv_seperator)

                if ids in table_page:
                    table_page[ids].append(table)
                else:
                    table_page[ids] = [table]
                # Extract table data and remaining content
                pattern = re.compile(r'<table>(.*?)(</table>)', re.DOTALL)
                data = item
                table_match = re.search(pattern, data)
                table_data = table_match.group(1) if table_match else ''
                remaining_content = data[table_match.end():] if table_match else data
                content[
                    idx] = f"<<table>><table>{table}</table><</table>>"  ## attach xml tags to differentiate table from other text
                count += 1
                # Check for list items in remaining content
                if "<<list>>" in remaining_content:
                    output = split_list_items_(remaining_content)
                    output = [x.strip() for x in output if x.strip()]
                    document_holder[ids].extend([content[idx]] + output)
                else:
                    document_holder[ids].extend([content[idx]] + [x.strip() for x in remaining_content.split('\n') if
                                                                  x.strip()])  # split other text by new line to be independent items in the python list.
            else:
                # Check for list items and tables in remaining content
                if "<<list>>" in item and "<table>" not in item:
                    output = split_list_items_(item)
                    output = [x.strip() for x in output if x.strip()]
                    document_holder[ids].extend(output)
                else:
                    document_holder[ids].extend([x.strip() for x in item.split("\n") if x.strip()])

    # # Flatten the nested list document_holder into a single list and Join the flattened list by "\n"
    flattened_list = [item for sublist in document_holder.values() for item in sublist]
    result = "\n".join(flattened_list)
    header_split = result.split("<titles>")

    def sub_header_content_splitta(string):
        """
        Splits the input string by XML tags and returns a list containing the segments of text,
        excluding segments containing specific XML tags such as "<header>", "<list>", or "<table>".

        Parameters:
        string (str): The input string to be processed.

        Returns:
        list: A list containing the segments of text extracted from the input string.
        """
        pattern = re.compile(r'<<[^>]+>>')
        segments = re.split(pattern, string)
        result = []
        for segment in segments:
            if segment.strip():
                if "<header>" not in segment and "<list>" not in segment and "<table>" not in segment:
                    segment = [x.strip() for x in segment.split('\n') if x.strip()]
                    result.extend(segment)
                else:
                    result.append(segment)
        return result

    max_words = 200
    chunks = {}
    table_header_dict = {}
    chunk_header_mapping = {}
    list_header_dict = {}

    # iterate through each title section
    for title_ids, items in enumerate(header_split):
        title_chunks = []
        current_chunk = []
        num_words = 0
        table_header_dict[title_ids] = {}
        chunk_header_mapping[title_ids] = {}
        list_header_dict[title_ids] = {}
        chunk_counter = 0
        for item_ids, item in enumerate(items.split('<headers>')):  # headers
            #print("".join(current_chunk).strip())
            lines = sub_header_content_splitta(item)
            SECTION_HEADER = None
            TITLES = None
            num_words = 0
            for ids_line, line in enumerate(lines):  # header lines

                if line.strip():
                    if "<title>" in line:
                        TITLES = re.findall(r'<title>(.*?)</title>', line)[0].strip()
                        line = TITLES
                        if re.sub(r'<[^>]+>', '', "".join(lines)).strip() == TITLES:
                            chunk_header_mapping[title_ids][chunk_counter] = lines
                            chunk_counter += 1
                    if "<header>" in line:
                        SECTION_HEADER = re.findall(r'<header>(.*?)</header>', line)[0].strip()
                        line = SECTION_HEADER
                        first_header_portion = True
                    next_num_words = num_words + len(re.findall(r'\w+', line))

                    if "<table>" not in line and "<list>" not in line:
                        if next_num_words > max_words and "".join(
                                current_chunk).strip() != SECTION_HEADER and current_chunk and "".join(
                                current_chunk).strip() != TITLES:

                            if SECTION_HEADER:
                                if first_header_portion:
                                    first_header_portion = False
                                else:
                                    current_chunk.insert(0, SECTION_HEADER.strip())

                            title_chunks.append(current_chunk)
                            chunk_header_mapping[title_ids][chunk_counter] = lines

                            current_chunk = []
                            num_words = 0
                            chunk_counter += 1

                        current_chunk.append(line)
                        num_words += len(re.findall(r'\w+', line))

                    """
                    Goal is to segment out table items and chunks intelligently.
                    We chunk the table by rows and for each chunk of the table we append the table column headers
                    and table headers if any. This way we preserve the table information across each chunks.
                    This will help improve semantic search where all the chunks relating to a table would be in the 
                    top k=n response giving the LLM mcomplet information on the table.
                    """

                    if "<table>" in line:
                        # Get table header which is usually line before table in document
                        line_index = lines.index(line)
                        if line_index != 0 and "<table>" not in lines[line_index - 1] and "<list>" not in lines[
                            line_index - 1]:  # Check if table is first item on the page, then they wont be a header (header may be included it table) and also if table is the the last item in the list
                            header = lines[line_index - 1].replace("<header>", "").replace("</header>", "")
                        else:
                            header = ""

                        table = line.split("<table>")[-1].split("</table>")[
                            0]  # get table from demarcators
                        df = pd.read_csv(io.StringIO(table), sep=csv_seperator, keep_default_na=False, header=None)
                        df.columns = df.iloc[0]
                        df = df[1:]
                        df.rename(columns=lambda x: '' if str(x).startswith('Unnamed:') else x, inplace=True)
                        table_chunks = []
                        curr_chunk = [df.columns.to_list()]  # start current chunk with table column names
                        words = len(re.findall(r'\w+', str(current_chunk) + " " + str(curr_chunk)))
                        # Iterate through the rows in the table
                        for row in df.itertuples(index=False):
                            curr_chunk.append(row)
                            words += len(re.findall(r'\w+', str(row)))
                            if words > max_words:
                                if [x for x in table_header_dict[title_ids] if chunk_counter == x]:
                                    table_header_dict[title_ids][chunk_counter].extend([header] + [table])
                                else:
                                    table_header_dict[title_ids][chunk_counter] = [header] + [table]
                                table_chunks.append("\n".join([csv_seperator.join(str(x) for x in curr_chunk[0])] + [
                                    csv_seperator.join(str(x) for x in r) for r in
                                    curr_chunk[1:]]))  # join chunk lines together to for a csv
                                tab_chunk = "\n".join([csv_seperator.join(str(x) for x in curr_chunk[0])] + [
                                    csv_seperator.join(str(x) for x in r) for r in
                                    curr_chunk[1:]])  # join chunk lines together to for a csv
                                words = len(re.findall(r'\w+', str(
                                    curr_chunk[0])))  # set word count to word length of column header names
                                if header:  # If header  attach header to table
                                    if current_chunk and current_chunk[
                                        -1].strip().lower() == header.strip().lower():  # check if header is in the chunk and remove to avoid duplicacy of header in chunk
                                        current_chunk.pop(-1)
                                    # Append section header to table
                                    if SECTION_HEADER and SECTION_HEADER.lower().strip() != header.lower().strip():
                                        if first_header_portion:
                                            first_header_portion = False
                                        else:
                                            current_chunk.insert(0, SECTION_HEADER.strip())
                                    current_chunk.extend([header.strip() + ':' if not header.strip().endswith(
                                        ':') else header.strip()] + [tab_chunk])  # enrich table header with ':'
                                    title_chunks.append(current_chunk)

                                else:
                                    if SECTION_HEADER:
                                        if first_header_portion:
                                            first_header_portion = False
                                        else:
                                            current_chunk.insert(0, SECTION_HEADER.strip())
                                    current_chunk.extend([tab_chunk])
                                    title_chunks.append(current_chunk)
                                chunk_header_mapping[title_ids][chunk_counter] = lines
                                chunk_counter += 1
                                num_words = 0
                                current_chunk = []
                                curr_chunk = [curr_chunk[0]]

                        if curr_chunk != [df.columns.to_list()] and lines.index(line) == len(
                                lines) - 1:  # if table chunk still remaining and table is last item in page append as last chunk
                            table_chunks.append("\n".join([csv_seperator.join(str(x) for x in curr_chunk[0])] + [
                                csv_seperator.join(str(x) for x in r) for r in curr_chunk[1:]]))
                            tab_chunk = "\n".join([csv_seperator.join(str(x) for x in curr_chunk[0])] + [
                                csv_seperator.join(str(x) for x in r) for r in curr_chunk[1:]])
                            if [x for x in table_header_dict[title_ids] if chunk_counter == x]:
                                table_header_dict[title_ids][chunk_counter].extend([header] + [table])
                            else:
                                table_header_dict[title_ids][chunk_counter] = [header] + [table]

                            if header:
                                if current_chunk and current_chunk[
                                    -1].strip().lower() == header.strip().lower():  # check if header is in the chunk and remove to avoid duplicacy of header in chunk
                                    current_chunk.pop(-1)
                                if SECTION_HEADER and SECTION_HEADER.lower().strip() != header.lower().strip():
                                    if first_header_portion:
                                        first_header_portion = False
                                    else:
                                        current_chunk.insert(0, SECTION_HEADER.strip())
                                current_chunk.extend(
                                    [header.strip() + ':' if not header.strip().endswith(':') else header.strip()] + [
                                        tab_chunk])
                                title_chunks.append(current_chunk)
                            else:
                                if SECTION_HEADER:
                                    if first_header_portion:
                                        first_header_portion = False
                                    else:
                                        current_chunk.insert(0, SECTION_HEADER.strip())
                                current_chunk.extend([tab_chunk])
                                title_chunks.append(current_chunk)
                            chunk_header_mapping[title_ids][chunk_counter] = lines
                            chunk_counter += 1
                            num_words = 0
                            current_chunk = []
                        elif curr_chunk != [df.columns.to_list()] and lines.index(line) != len(
                                lines) - 1:  # if table is not last item in page and max word threshold is not reached, send no next loop
                            table_chunks.append("\n".join([csv_seperator.join(str(x) for x in curr_chunk[0])] + [
                                csv_seperator.join(str(x) for x in r) for r in curr_chunk[1:]]))
                            tab_chunk = "\n".join([csv_seperator.join(str(x) for x in curr_chunk[0])] + [
                                csv_seperator.join(str(x) for x in r) for r in curr_chunk[1:]])

                            if [x for x in table_header_dict[title_ids] if chunk_counter == x]:
                                table_header_dict[title_ids][chunk_counter].extend([header] + [table])
                            else:
                                table_header_dict[title_ids][chunk_counter] = [header] + [table]
                            if header:
                                if current_chunk and current_chunk[
                                    -1].strip().lower() == header.strip().lower():  # check if header is in the chunk and remove to avoid duplicacy of header in chunk
                                    current_chunk.pop(-1)
                                current_chunk.extend(
                                    [header.strip() + ':' if not header.strip().endswith(':') else header.strip()] + [
                                        tab_chunk])
                            else:
                                current_chunk.extend([tab_chunk])
                            num_words = words

                    """
                    Goal is to segment out list items and chunk intelligently.
                    We chunk each list by items in the list and 
                    for each list chunk we append the list header to the chunk to preserve the information of the list across chunks.
                    This would boost retrieval process where question pertaining to a list will have all list chunks within
                    the topK=n responses.
                    """

                    if "<list>" in line:
                        # Get list header which is usually line before list in document
                        line_index = lines.index(line)
                        if line_index != 0 and "<table>" not in lines[line_index - 1] and "<list>" not in lines[
                            line_index - 1]:  # Check if table or list is the previous item on the page, then they wont be a header
                            header = lines[line_index - 1].replace("<header>", "").replace("</header>", "")
                        else:
                            header = ""
                        list_pattern = re.compile(r'<list>(.*?)(?:</list>|$)',
                                                  re.DOTALL)  ## Grab all list contents within the list xml tags
                        list_match = re.search(list_pattern, line)
                        list_ = list_match.group(1)
                        list_lines = list_.split("\n")

                        curr_chunk = []
                        words = len(re.findall(r'\w+', str(current_chunk)))  # start word count from any existing chunk
                        # Iterate through the items in the list
                        for lyst_item in list_lines:
                            curr_chunk.append(lyst_item)
                            words += len(re.findall(r'\w+', lyst_item))
                            if words >= max_words:  #
                                if [x for x in list_header_dict[title_ids] if chunk_counter == x]:
                                    list_header_dict[title_ids][chunk_counter].extend([header] + [list_])
                                else:
                                    list_header_dict[title_ids][chunk_counter] = [header] + [list_]
                                words = 0
                                list_chunk = "\n".join(curr_chunk)
                                if header:  # attach list header
                                    if current_chunk and current_chunk[
                                        -1].strip().lower() == header.strip().lower():  # check if header is in the chunk and remove to avoid duplicacy of header in chunk
                                        current_chunk.pop(-1)
                                        # Append section content header to list
                                    if SECTION_HEADER and SECTION_HEADER.lower().strip() != header.lower().strip():
                                        if first_header_portion:
                                            first_header_portion = False
                                        else:
                                            current_chunk.insert(0, SECTION_HEADER.strip())

                                    current_chunk.extend([header.strip() + ':' if not header.strip().endswith(
                                        ':') else header.strip()] + [list_chunk])
                                    title_chunks.append(current_chunk)

                                else:
                                    if SECTION_HEADER:
                                        if first_header_portion:
                                            first_header_portion = False
                                        else:
                                            current_chunk.insert(0, SECTION_HEADER.strip())

                                    current_chunk.extend([list_chunk])
                                    title_chunks.append(current_chunk)
                                chunk_header_mapping[title_ids][chunk_counter] = lines
                                chunk_counter += 1
                                num_words = 0
                                current_chunk = []
                                curr_chunk = []
                        if curr_chunk and lines.index(line) == len(
                                lines) - 1:  # if list chunk still remaining and list is last item in page append as last chunk
                            list_chunk = "\n".join(curr_chunk)
                            if [x for x in list_header_dict[title_ids] if chunk_counter == x]:
                                list_header_dict[title_ids][chunk_counter].extend([header] + [list_])
                            else:
                                list_header_dict[title_ids][chunk_counter] = [header] + [list_]
                            if header:
                                if current_chunk and current_chunk[
                                    -1].strip().lower() == header.strip().lower():  # check if header is in the chunk and remove to avoid duplicacy of header in chunk
                                    current_chunk.pop(-1)
                                if SECTION_HEADER and SECTION_HEADER.lower().strip() != header.lower().strip():
                                    if first_header_portion:
                                        first_header_portion = False
                                    else:
                                        current_chunk.insert(0, SECTION_HEADER.strip())
                                current_chunk.extend(
                                    [header.strip() + ':' if not header.strip().endswith(':') else header.strip()] + [
                                        list_chunk])
                                title_chunks.append(current_chunk)
                            else:
                                if SECTION_HEADER:
                                    if first_header_portion:
                                        first_header_portion = False
                                    else:
                                        current_chunk.insert(0, SECTION_HEADER.strip())
                                current_chunk.extend([list_chunk])
                                title_chunks.append(current_chunk)
                            chunk_header_mapping[title_ids][chunk_counter] = lines
                            chunk_counter += 1
                            num_words = 0
                            current_chunk = []
                        elif curr_chunk and lines.index(line) != len(
                                lines) - 1:  # if list is not last item in page and max word threshold is not reached, send to next loop
                            list_chunk = "\n".join(curr_chunk)
                            if [x for x in list_header_dict[title_ids] if chunk_counter == x]:
                                list_header_dict[title_ids][chunk_counter].extend([header] + [list_])
                            else:
                                list_header_dict[title_ids][chunk_counter] = [header] + [list_]
                            if header:
                                if current_chunk and current_chunk[
                                    -1].strip().lower() == header.strip().lower():  # check if header is in the chunk and remove to avoid duplicacy of header in chunk
                                    current_chunk.pop(-1)
                                current_chunk.extend(
                                    [header.strip() + ':' if not header.strip().endswith(':') else header.strip()] + [
                                        list_chunk])
                            else:
                                current_chunk.extend([list_chunk])
                            num_words = words

            if current_chunk and "".join(current_chunk).strip() != SECTION_HEADER and "".join(
                    current_chunk).strip() != TITLES:

                if SECTION_HEADER:
                    if first_header_portion:
                        first_header_portion = False
                    else:
                        current_chunk.insert(0, SECTION_HEADER.strip())
                title_chunks.append(current_chunk)
                chunk_header_mapping[title_ids][chunk_counter] = lines
                current_chunk = []
                chunk_counter += 1
        if current_chunk:
            title_chunks.append(current_chunk)
            chunk_header_mapping[title_ids][chunk_counter] = lines
        chunks[title_ids] = title_chunks

    # List of title header sections document was split into
    for x in chunk_header_mapping:
        if chunk_header_mapping[x]:
            try:
                title_pattern = re.compile(r'<title>(.*?)(?:</title>|$)', re.DOTALL)
                title_match = re.search(title_pattern, chunk_header_mapping[x][0][0])
                title_ = title_match.group(1) if title_match else ""
                #print(title_, end='\n')
            except:
                continue

    with open(f"json_data/{doc_id}.json", "w") as f:
        json.dump(chunk_header_mapping, f)
    s3.upload_file(f"json_data/{doc_id}.json", BUCKET, f"{doc_id}.json")

    from opensearchpy import OpenSearch, RequestsHttpConnection, AWSV4SignerAuth
    from requests_aws4auth import AWS4Auth

    """
    This script demonstrates indexing documents into an Amazon OpenSearch Service domain using AWS Identity and Access Management (IAM) for authentication.
    """
    # Embedding Model
    model = "titanv2"
    # Use OpenSearch Servelss or Not
    YOUR_OPENSEARCH_ENDPOINT = True
    service = 'aoss'
    # replace wit your OpenSearch Service domain/Serverless endpoint
    domain_endpoint = "qlocxlzg30mhnqbqq628.us-west-2.aoss.amazonaws.com"

    credentials = boto3.Session().get_credentials()
    awsauth = AWSV4SignerAuth(credentials, "us-west-2", service)
    os_ = OpenSearch(
        hosts=[{'host': domain_endpoint, 'port': 443}],
        http_auth=awsauth,
        use_ssl=True,
        verify_certs=True,
        timeout=300,
        # http_compress = True, # enables gzip compression for request bodies
        connection_class=RequestsHttpConnection
    )

    # Sample Opensearch domain index mapping
    # mapping = {
    #     'settings': {
    #         'index': {
    #             'knn': True,
    #             "knn.algo_param.ef_search": 100,
    #         }
    #     },

    #     'mappings': {
    #         'properties': {
    #             'embedding': {
    #                 'type': 'knn_vector',
    #                 'dimension': model_dimension_mapping[model],  # change as per sequence length of Embedding Model
    #                 "method": {
    #                     "name": "hnsw",
    #                     "space_type": "cosinesimil",
    #                     "engine": "nmslib",
    #                     "parameters": {
    #                         "ef_construction": 256,
    #                         "m": 48
    #                     }
    #                 }
    #             },

    #             'passage': {
    #                 'type': 'text'
    #             },

    #             'page': {
    #                 'type': 'long'
    #             },

    #             'doc_id': {
    #                 'type': 'keyword'
    #             },

    #             'table': {
    #                 'type': 'text'
    #             },

    #             'list': {
    #                 'type': 'text'
    #             },

    #             'title_headers': {
    #                 'type': 'text'
    #             },
    #             'section_header_ids': {
    #                 'type': 'text'
    #             },
    #             'section_title_ids': {
    #                 'type': 'text'
    #             },
    #             'url': {
    #                 'type': 'text'
    #             },

    #         }
    #     }
    # }

    #domain_index = f"test2-{model}-new"  # domain index name

    domain_index = INDEX_NAME

    # Index creation code
    # if not os_.indices.exists(index=domain_index):
    #     print("Index not exists")
    #     os_.indices.create(index=domain_index, body=mapping)
    #     # Verify that the index has been created
    #     if os_.indices.exists(index=domain_index):
    #         print(f"Index {domain_index} created successfully.")
    #     else:
    #         print(f"Failed to create index '{domain_index}'.")
    # else:
    #     print(f'{domain_index} Index already exists!')

    print("Inserting document into index... ")
    #print(chunks)
    i = 1
    SAGEMAKER = boto3.client('sagemaker-runtime')
    page_text = get_pdf_stream(document_url, "PRISMresource")
    for ids, chunkks in chunks.items():  # Iterate through the page title chunks
        try:
            index_adjuster = len(chunk_header_mapping[ids]) % len(chunkks)
        except ZeroDivisionError:
            print(f"Skipping title ID: {ids} due to empty chunks.")
            continue
        for chunk_ids, chunk in enumerate(chunkks):  # iterating through section header chunks
            chunk_ids += index_adjuster
            passage_chunk = "\n".join(chunk).replace("<title>", "").replace("</title>", "")
            if passage_chunk.strip():
                #print(passage_chunk)
                embedding = _get_emb_(passage_chunk, model)
                table = []
                if ids in table_header_dict:
                    if [x for x in table_header_dict[ids] if x == chunk_ids]:
                        table = "\n".join(table_header_dict[ids][chunk_ids])
                lists = []
                if ids in list_header_dict:
                    if [x for x in list_header_dict[ids] if x == chunk_ids]:
                        lists = "\n".join(list_header_dict[ids][chunk_ids])

                page_num = get_page_num(page_text, passage_chunk)
                documentt = {
                    'doc_id': doc_id,  # doc name
                    'passage': passage_chunk,
                    'page': page_num,
                    'embedding': embedding,
                    'table': table,
                    "list": lists,
                    "section_header_ids": chunk_ids,  # Store id of the header section
                    "section_title_ids": ids,  # Store id of the title section
                    "url": document_url,
                }

                try:
                    response = os_.index(index=domain_index, body=documentt)
                    i += 1
                    # Check the response to see if the indexing was successful
                    if response["result"] == "created":
                        print(f"Document indexed successfully with ID: {response['_id']}")
                    else:
                        print("Failed to index document.")
                except RequestError as e:
                    logging.error(f"Error indexing document to index '{domain_index}': {e}")
            else:
                continue


def process_xlsx(bucket_name, file_name, text, url):
    from opensearchpy import OpenSearch, RequestsHttpConnection, AWSV4SignerAuth
    from requests_aws4auth import AWS4Auth
    # Embedding Model
    model = "titanv2"
    # Use OpenSearch Servelss or Not
    YOUR_OPENSEARCH_ENDPOINT = True
    service = 'aoss'
    # replace wit your OpenSearch Service domain/Serverless endpoint
    domain_endpoint = "qlocxlzg30mhnqbqq628.us-west-2.aoss.amazonaws.com"
    
    credentials = boto3.Session().get_credentials()
    awsauth = AWSV4SignerAuth(credentials, "us-west-2", service)
    os_ = OpenSearch(
        hosts=[{'host': domain_endpoint, 'port': 443}],
        http_auth=awsauth,
        use_ssl=True,
        verify_certs=True,
        timeout=300,
        # http_compress = True, # enables gzip compression for request bodies
        connection_class=RequestsHttpConnection
    )

    def insert_document(doc):
        response = os_.index(
            index=INDEX_NAME,
            body=doc
        )
        return response

    document = {
        'doc_id': file_name,
        'passage': text,
        'page': None,
        'embedding': _get_emb_(text, model),
        'table': None,
        'list': None,
        'section_header_ids': None,
        'section_title_ids': None,
        'url': url
    }

    response = insert_document(document)


#bucket_name = "YOUR_BUCKET_NAME"

#file_name = "data/vulnerability-management-best-practices.pdf"
#file_name = "/home/ec2-user/prism/data_processing/vulner.pdf"
# file_name = "dot-drug-and-alcohol-program-missed-tests"  # Change to file path either in S3 or Local

#process_document(bucket_name, file_name, "urlll")
